from pyexpat.errors import messages
from django.shortcuts import get_object_or_404, render
from django.shortcuts import render, redirect
# Create your views here.
from django.contrib.auth import login
from ecom.models import ShowroomStaff
from .forms import ShowroomStaffUpdateForm, UserUpdateForm
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages

def showroomdash(request):
    return render(request,'Showroomstaff_dash.html')
@login_required
def update_showroom_staff(request):
    showroom_staff = ShowroomStaff.objects.get(userid=request.user)

    if request.method == 'POST':
        user_form = UserUpdateForm(request.POST, instance=request.user)
        showroom_form = ShowroomStaffUpdateForm(request.POST, instance=showroom_staff)

        if user_form.is_valid() and showroom_form.is_valid():
            user_form.save()
            showroom_form.save()
            messages.success(request, 'Profile updated successfully!')
            return redirect('showroom_staff_profile_update')  # Redirect to the same update page
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        user_form = UserUpdateForm(instance=request.user)
        showroom_form = ShowroomStaffUpdateForm(instance=showroom_staff)

    return render(request, 'showroom_profile_update.html', {
        'user_form': user_form,
        'showroom_form': showroom_form,
    })


from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa  # For PDF generation
from ecom.models import Customer, FieldStaff, tbl_product, Order, OrderItem
import json
from datetime import date, datetime
import uuid
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.utils.timezone import now
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
import uuid

from ecom.models import Customer, Order, OrderItem, tbl_product, FieldStaff
from django.contrib.auth.models import User

from django.contrib import messages
from django.http import JsonResponse
from django.utils.timezone import now
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.utils.timezone import now
from django.contrib.auth.models import User

from ecom.models import Customer, FieldStaff, tbl_product, Order, OrderItem, ShowroomStaff
from .forms import ShowroomStaffUpdateForm, UserUpdateForm
from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count
from django.utils import timezone
from datetime import timedelta
import csv
import io
import xlsxwriter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from openpyxl import Workbook # type: ignore


# Adjust import based on your app structure
'''
@login_required
def inventory_dashboard(request):
    """
    View to display sales information dashboard with charts and reports.
    """
    # Get filter parameters
    date_range = request.GET.get('date-range', '30')  # Default to 30 days
    category_id = request.GET.get('product-category', 'all')
    
    # Set date filter
    today = timezone.now().date()
    if date_range == 'custom':
        start_date = request.GET.get('start-date')
        end_date = request.GET.get('end-date')
        if start_date and end_date:
            try:
                start_date = timezone.datetime.strptime(start_date, '%Y-%m-%d').date()
                end_date = timezone.datetime.strptime(end_date, '%Y-%m-%d').date()
            except ValueError:
                start_date = today - timedelta(days=30)
                end_date = today
        else:
            start_date = today - timedelta(days=30)
            end_date = today
    else:
        try:
            days = int(date_range)
            start_date = today - timedelta(days=days)
            end_date = today
        except ValueError:
            start_date = today - timedelta(days=30)
            end_date = today
    
    # Filter orders by date
    orders = Order.objects.filter(date__gte=start_date, date__lte=end_date)
    
    # Filter by category if specified
    if category_id != 'all':
        try:
            category = tbl_category.objects.get(id=int(category_id))
            order_items = OrderItem.objects.filter(
                order__in=orders,
                product__category=category
            )
            orders = Order.objects.filter(id__in=order_items.values_list('order_id', flat=True))
        except (ValueError, tbl_category.DoesNotExist):
            pass
    
    # Calculate summary metrics
    total_orders = orders.count()
    total_revenue = orders.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    # Get units sold
    order_items = OrderItem.objects.filter(order__in=orders)
    total_units_sold = order_items.aggregate(Sum('quantity'))['quantity__sum'] or 0
    
    # Get unique customers
    unique_customers = orders.values('user').distinct().count()
    
    # Get sales trend data
    sales_data = []
    for i in range((end_date - start_date).days + 1):
        current_date = start_date + timedelta(days=i)
        daily_sales = orders.filter(date=current_date).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        sales_data.append((current_date, daily_sales))
    
    # Get best selling products
    best_selling_products = (
        OrderItem.objects.filter(order__in=orders)
        .values('product')
        .annotate(sold_count=Sum('quantity'))
        .order_by('-sold_count')[:5]
    )
    
    # Fetch product details
    for product in best_selling_products:
        try:
            prod = tbl_product.objects.get(id=product['product'])
            product['product_name'] = prod.product_name
        except tbl_product.DoesNotExist:
            product['product_name'] = 'Unknown Product'
    
    # Get sales by category
    category_sales = []
    categories = tbl_category.objects.all()
    for category in categories:
        category_orders = OrderItem.objects.filter(
            order__in=orders,
            product__category=category
        )
        category_amount = category_orders.aggregate(
            total=Sum('subtotal')
        )['total'] or 0
        
        if category_amount > 0:
            category_sales.append({
                'name': category.category_name,
                'amount': category_amount
            })
    
    # Sort categories by sales amount
    category_sales = sorted(category_sales, key=lambda x: x['amount'], reverse=True)
    
    # Get recent orders for the table
    recent_orders = orders.order_by('-date')[:20]
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'total_orders': total_orders,
        'total_revenue': total_revenue,
        'total_units_sold': total_units_sold,
        'unique_customers': unique_customers,
        'sales_dates': [data[0] for data in sales_data],
        'sales_amounts': [data[1] for data in sales_data],
        'best_selling_products': best_selling_products,
        'category_sales': category_sales,
        'recent_orders': recent_orders,
        'categories': categories,
    }
    
    return render(request, 'inventorydash.html', context)'''

from django.shortcuts import redirect
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta
from django.db.models import Sum
from django.contrib.auth.decorators import login_required
from fpdf import FPDF
from openpyxl import Workbook
import os
from django.conf import settings

from ecom.models import Order, OrderItem, tbl_category, tbl_product

'''
@login_required
def download_report(request):
    """
    Generate and download sales report in PDF or Excel format.
    """
    # Get filter parameters
    date_range = request.GET.get('date-range', '30')
    category_id = request.GET.get('product-category', 'all')
    report_format = request.GET.get('format', 'pdf')
    
    today = timezone.now().date()
    if date_range == 'custom':
        start_date_str = request.GET.get('start-date')
        end_date_str = request.GET.get('end-date')
        try:
            start_date = timezone.datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = timezone.datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except (TypeError, ValueError):
            start_date = today - timedelta(days=30)
            end_date = today
    else:
        try:
            days = int(date_range)
            start_date = today - timedelta(days=days)
            end_date = today
        except ValueError:
            start_date = today - timedelta(days=30)
            end_date = today
    
    # Filter orders
    orders = Order.objects.filter(date__gte=start_date, date__lte=end_date)
    
    if category_id != 'all':
        try:
            category = tbl_category.objects.get(id=int(category_id))
            order_items = OrderItem.objects.filter(
                order__in=orders,
                product__category=category
            )
            orders = Order.objects.filter(id__in=order_items.values_list('order_id', flat=True))
        except (ValueError, tbl_category.DoesNotExist):
            pass

    filename = f"sales_report_{start_date.strftime('%Y%m%d')}_to_{end_date.strftime('%Y%m%d')}"

    if report_format == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'

        pdf = FPDF()
        pdf.add_page()

        # Use built-in helvetica font which supports bold
        pdf.set_font('helvetica', 'B', 16)

        # Header
        pdf.cell(0, 10, 'Sales Report', 0, 1, 'C')
        pdf.set_font('helvetica', '', 12)
        pdf.cell(0, 10, f'Period: {start_date} to {end_date}', 0, 1, 'C')
        pdf.ln(10)

        # Summary
        total_orders = orders.count()
        total_revenue = orders.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        order_items = OrderItem.objects.filter(order__in=orders)
        total_units_sold = order_items.aggregate(Sum('quantity'))['quantity__sum'] or 0
        unique_customers = orders.values('user').distinct().count()

        pdf.set_font('helvetica', 'B', 14)
        pdf.cell(0, 10, 'Summary', 0, 1)
        pdf.set_font('helvetica', '', 12)
        pdf.cell(60, 10, 'Total Orders:', 0)
        pdf.cell(0, 10, str(total_orders), 0, 1)
        pdf.cell(60, 10, 'Total Revenue:', 0)
        pdf.cell(0, 10, f'Rs. {total_revenue}', 0, 1)
        pdf.cell(60, 10, 'Units Sold:', 0)
        pdf.cell(0, 10, str(total_units_sold), 0, 1)
        pdf.cell(60, 10, 'Unique Customers:', 0)
        pdf.cell(0, 10, str(unique_customers), 0, 1)
        pdf.ln(10)

        # Order Details
        pdf.set_font('helvetica', 'B', 14)
        pdf.cell(0, 10, 'Order Details', 0, 1)
        pdf.set_font('helvetica', 'B', 12)
        pdf.cell(30, 10, 'Order ID', 1)
        pdf.cell(30, 10, 'Date', 1)
        pdf.cell(50, 10, 'Customer', 1)
        pdf.cell(30, 10, 'Amount', 1)
        pdf.cell(40, 10, 'Status', 1)
        pdf.ln()

        pdf.set_font('helvetica', '', 12)
        for order in orders:
            # Make sure text doesn't overflow cells by limiting their width
            order_id = str(order.invoice_no)
            date = order.date.strftime('%Y-%m-%d')
            username = order.user.username
            if len(username) > 20:  # Truncate long usernames
                username = username[:17] + '...'
            amount = f'₹{order.total_amount}'
            status = order.status

            pdf.cell(30, 10, order_id, 1)
            pdf.cell(30, 10, date, 1)
            pdf.cell(50, 10, username, 1)
            pdf.cell(30, 10, amount, 1)
            pdf.cell(40, 10, status, 1)
            pdf.ln()

        # Handle potential encoding issues safely
        try:
            response.write(pdf.output(dest='S').encode('latin1'))
        except UnicodeEncodeError:
            # Fallback for characters not in latin1 charset
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font('helvetica', 'B', 16)
            pdf.cell(0, 10, 'Sales Report', 0, 1, 'C')
            pdf.set_font('helvetica', '', 12)
            pdf.cell(0, 10, 'Error: Some characters could not be encoded. Please use Excel format instead.', 0, 1)
            response.write(pdf.output(dest='S').encode('latin1'))
        
        return response

    elif report_format == 'excel':
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'

        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Report"

        ws.append(['Sales Report'])
        ws.append([f'Period: {start_date} to {end_date}'])
        ws.append([])

        # Summary
        total_orders = orders.count()
        total_revenue = orders.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        order_items = OrderItem.objects.filter(order__in=orders)
        total_units_sold = order_items.aggregate(Sum('quantity'))['quantity__sum'] or 0
        unique_customers = orders.values('user').distinct().count()

        ws.append(['Summary'])
        ws.append(['Total Orders:', total_orders])
        ws.append(['Total Revenue:', f'₹{total_revenue}'])
        ws.append(['Units Sold:', total_units_sold])
        ws.append(['Unique Customers:', unique_customers])
        ws.append([])

        # Best-selling products
        ws.append(['Best Selling Products'])
        ws.append(['Product', 'Units Sold'])
        best_selling = (
            OrderItem.objects.filter(order__in=orders)
            .values('product')
            .annotate(sold_count=Sum('quantity'))
            .order_by('-sold_count')[:10]
        )
        for item in best_selling:
            try:
                prod = tbl_product.objects.get(id=item['product'])
                name = prod.product_name
            except tbl_product.DoesNotExist:
                name = 'Unknown Product'
            ws.append([name, item['sold_count']])
        ws.append([])

        # Order details
        ws.append(['Order Details'])
        ws.append(['Order ID', 'Date', 'Customer', 'Products', 'Amount', 'Status'])
        for order in orders:
            items = OrderItem.objects.filter(order=order)
            product_str = ', '.join([f"{i.quantity} x {i.product.product_name}" for i in items])
            ws.append([
                order.invoice_no,
                order.date.strftime('%Y-%m-%d'),
                order.user.username,
                product_str,
                f'₹{order.total_amount}',
                order.status
            ])

        wb.save(response)
        return response

    else:
        return redirect('inventory_dashboard')'''
from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.utils import timezone
from datetime import timedelta
from openpyxl import Workbook

# Using ReportLab instead of FPDF
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from io import BytesIO

from ecom.models import Order, OrderItem, tbl_product, tbl_category
@login_required
def inventory_dashboard(request):
    """
    View to display sales information dashboard with charts and reports.
    """
    # Get filter parameters with proper defaults
    date_range = request.GET.get('date-range', '30')  # Default to 30 days
    category_id = request.GET.get('product-category', 'all')
    
    # Set date filter
    today = timezone.now().date()
    
    # Handle date range filtering
    if date_range == 'custom':
        try:
            start_date_str = request.GET.get('start-date')
            end_date_str = request.GET.get('end-date')
            if start_date_str and end_date_str:
                start_date = timezone.datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = timezone.datetime.strptime(end_date_str, '%Y-%m-%d').date()
            else:
                # Default to 30 days if custom is selected but dates aren't provided
                start_date = today - timedelta(days=30)
                end_date = today
        except ValueError:
            # Handle invalid date format
            start_date = today - timedelta(days=30)
            end_date = today
    else:
        # Handle predefined ranges (7, 30, 90, 180, 365 days)
        try:
            days = int(date_range)
            start_date = today - timedelta(days=days)
            end_date = today
        except ValueError:
            # If conversion fails, default to 30 days
            start_date = today - timedelta(days=30)
            end_date = today
    
    # Ensure end_date is not in the future
    if end_date > today:
        end_date = today
    
    # Ensure start_date is not after end_date
    if start_date > end_date:
        start_date = end_date - timedelta(days=30)
    
    # Get all orders within date range
    # Include the full end date by adding a day and subtracting 1 second
    orders = Order.objects.filter(
        date__gte=start_date, 
        date__lte=end_date + timedelta(days=1) - timedelta(seconds=1)
    )
    
    # Apply category filter if specified
    if category_id != 'all' and category_id:
        try:
            category_id = int(category_id)
            # Check if category exists
            category = tbl_category.objects.get(id=category_id)
            
            # Get order IDs that have products from this category
            category_order_ids = OrderItem.objects.filter(
                order__in=orders,
                product__category=category
            ).values_list('order_id', flat=True).distinct()
            
            # Filter orders by these IDs
            orders = orders.filter(id__in=category_order_ids)
        except (ValueError, tbl_category.DoesNotExist):
            # Invalid category ID, keep all orders
            pass
    
    # Calculate summary metrics
    total_orders = orders.count()
    total_revenue = orders.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    # Get units sold
    order_items = OrderItem.objects.filter(order__in=orders)
    total_units_sold = order_items.aggregate(Sum('quantity'))['quantity__sum'] or 0
    
    # Get unique customers
    unique_customers = orders.values('user').distinct().count()
    
    # Get sales trend data for each day in the range
    sales_data = []
    date_range_days = (end_date - start_date).days + 1
    
    # Create a dictionary to store daily sales
    daily_sales_dict = {}
    
    # Initialize all dates in the range with 0 sales
    current_date = start_date
    while current_date <= end_date:
        daily_sales_dict[current_date] = 0
        current_date += timedelta(days=1)
    
    # Get actual sales for each day
    daily_orders = orders.values('date').annotate(daily_total=Sum('total_amount'))
    
    # Fill in actual sales data
    for daily_order in daily_orders:
        order_date = daily_order['date'].date() if isinstance(daily_order['date'], datetime) else daily_order['date']
        if order_date in daily_sales_dict:
            daily_sales_dict[order_date] = daily_order['daily_total']
    
    # Convert to ordered list for template
    sales_dates = []
    sales_amounts = []
    for date_key in sorted(daily_sales_dict.keys()):
        sales_dates.append(date_key)
        sales_amounts.append(daily_sales_dict[date_key])
    
    # Get best selling products
    best_selling_products = []
    top_products = (
        OrderItem.objects.filter(order__in=orders)
        .values('product')
        .annotate(sold_count=Sum('quantity'))
        .order_by('-sold_count')[:5]
    )
    
    # Fetch product details
    for product in top_products:
        try:
            prod = tbl_product.objects.get(id=product['product'])
            best_selling_products.append({
                'product_name': prod.product_name,
                'sold_count': product['sold_count']
            })
        except tbl_product.DoesNotExist:
            best_selling_products.append({
                'product_name': 'Unknown Product',
                'sold_count': product['sold_count']
            })
    
    # Get sales by category
    category_sales = []
    categories = tbl_category.objects.all()
    for category in categories:
        category_items = OrderItem.objects.filter(
            order__in=orders,
            product__category=category
        )
        category_amount = category_items.aggregate(
            total=Sum('subtotal')
        )['total'] or 0
        
        if category_amount > 0:
            category_sales.append({
                'name': category.category_name,
                'amount': category_amount
            })
    
    # Sort categories by sales amount
    category_sales = sorted(category_sales, key=lambda x: x['amount'], reverse=True)
    
    # Get recent orders for the table
    recent_orders = orders.order_by('-date')[:20]
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'total_orders': total_orders,
        'total_revenue': total_revenue,
        'total_units_sold': total_units_sold,
        'unique_customers': unique_customers,
        'sales_dates': sales_dates,
        'sales_amounts': sales_amounts,
        'best_selling_products': best_selling_products,
        'category_sales': category_sales,
        'recent_orders': recent_orders,
        'categories': categories,
        # Add the current filter settings to persist in the form
        'selected_date_range': date_range,
        'selected_category': category_id,
    }
    
    return render(request, 'inventorydash.html', context)

@login_required
def download_report(request):
    """
    Generate and download sales report in PDF or Excel format with improved formatting.
    """
    # Get filter parameters
    date_range = request.GET.get('date-range', '30')
    category_id = request.GET.get('product-category', 'all')
    report_format = request.GET.get('format', 'pdf')
    
    # Set date filter
    today = timezone.now().date()
    
    # Handle custom date range
    if date_range == 'custom':
        try:
            start_date_str = request.GET.get('start-date')
            end_date_str = request.GET.get('end-date')
            if start_date_str and end_date_str:
                start_date = timezone.datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = timezone.datetime.strptime(end_date_str, '%Y-%m-%d').date()
            else:
                start_date = today - timedelta(days=30)
                end_date = today
        except ValueError:
            start_date = today - timedelta(days=30)
            end_date = today
    else:
        # Handle predefined ranges (7, 30, 90, 180, 365 days)
        try:
            days = int(date_range)
            start_date = today - timedelta(days=days)
            end_date = today
        except ValueError:
            start_date = today - timedelta(days=30)
            end_date = today
    
    # Ensure end_date is not in the future
    if end_date > today:
        end_date = today
    
    # Ensure start_date is not after end_date
    if start_date > end_date:
        start_date = end_date - timedelta(days=30)
    
    # Get all orders within date range
    # Important: Include the full end date by adding a day and subtracting a second
    orders = Order.objects.filter(
        date__gte=start_date,
        date__lte=end_date + timedelta(days=1) - timedelta(seconds=1)
    )
    
    # Apply category filter if specified
    if category_id != 'all' and category_id:
        try:
            category_id = int(category_id)
            # Check if category exists
            category = tbl_category.objects.get(id=category_id)
            
            # Get product IDs for this category
            category_product_ids = tbl_product.objects.filter(
                category=category
            ).values_list('id', flat=True)
            
            # Get order IDs that have products from this category
            category_order_ids = OrderItem.objects.filter(
                order__in=orders,
                product_id__in=category_product_ids
            ).values_list('order_id', flat=True).distinct()
            
            # Filter orders by these IDs
            orders = orders.filter(id__in=category_order_ids)
        except (ValueError, tbl_category.DoesNotExist):
            # Invalid category ID, keep all orders
            pass


    filename = f"sales_report_{start_date.strftime('%Y%m%d')}_to_{end_date.strftime('%Y%m%d')}"

    # Calculate summary metrics for report
    total_orders = orders.count()
    total_revenue = orders.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    order_items = OrderItem.objects.filter(order__in=orders)
    total_units_sold = order_items.aggregate(Sum('quantity'))['quantity__sum'] or 0
    unique_customers = orders.values('user').distinct().count()

    # Get best selling products for reports
    best_selling = (
        OrderItem.objects.filter(order__in=orders)
        .values('product')
        .annotate(sold_count=Sum('quantity'))
        .order_by('-sold_count')[:10]
    )
    
    best_selling_products = []
    for item in best_selling:
        try:
            prod = tbl_product.objects.get(id=item['product'])
            best_selling_products.append({
                'name': prod.product_name,
                'sold_count': item['sold_count']
            })
        except tbl_product.DoesNotExist:
            best_selling_products.append({
                'name': 'Unknown Product',
                'sold_count': item['sold_count']
            })

    # Generate category data for reports
    category_data = []
    categories = tbl_category.objects.all()
    for category in categories:
        category_items = OrderItem.objects.filter(
            order__in=orders,
            product__category=category
        )
        category_amount = category_items.aggregate(
            total=Sum('subtotal')
        )['total'] or 0
        
        if category_amount > 0:
            category_data.append({
                'name': category.category_name,
                'amount': category_amount
            })

    if report_format == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        
        # Create a buffer to receive PDF data
        buffer = BytesIO()
        
        try:
            # Create the PDF object, using the buffer as its "file"
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            
            # Container for the PDF elements
            elements = []
            
            # Get sample stylesheet
            styles = getSampleStyleSheet()
            title_style = styles['Heading1']
            subtitle_style = styles['Heading2']
            normal_style = styles['Normal']
            
            # Add the title
            elements.append(Paragraph('Sales Report', title_style))
            elements.append(Paragraph(f'Period: {start_date} to {end_date}', subtitle_style))
            elements.append(Spacer(1, 20))
            
            # Summary section
            elements.append(Paragraph('Summary', subtitle_style))
            summary_data = [
                ['Total Orders:', str(total_orders)],
                ['Total Revenue:', f'Rs.{total_revenue:.2f}'],
                ['Units Sold:', str(total_units_sold)],
                ['Unique Customers:', str(unique_customers)]
            ]
            
            summary_table = Table(summary_data, colWidths=[150, 350])
            summary_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 20))
            
            # Best Selling Products - with improved wrapping
            elements.append(Paragraph('Best Selling Products', subtitle_style))
            best_products_data = [['Product', 'Units Sold']]
            
            for product in best_selling_products:
                # Wrap long product names in Paragraph for proper wrapping
                name_paragraph = Paragraph(product['name'], normal_style)
                best_products_data.append([name_paragraph, str(product['sold_count'])])
            
            # Adjust column widths - give more space to product names
            best_products_table = Table(best_products_data, colWidths=[400, 100])
            best_products_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ALIGN', (1, 1), (1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('WORDWRAP', (0, 1), (0, -1), True),  # Enable word wrapping for product names
            ]))
            elements.append(best_products_table)
            elements.append(Spacer(1, 20))
            
            # Order Details with improved column alignment
            elements.append(Paragraph('Order Details', subtitle_style))
            
            order_data = [['Order ID', 'Date', 'Customer', 'Amount']]
            for order in orders[:50]:  # Limit to 50 orders

                customer_name = getattr(order, 'customer_name', order.user.get_full_name() or order.user.username)
                # Wrap username in Paragraph for proper wrapping
                name_paragraph = Paragraph(customer_name, normal_style)
                
                order_data.append([
                    str(order.invoice_no),
                    order.date.strftime('%Y-%m-%d'),
                    name_paragraph,
                    f'Rs.{order.total_amount:.2f}',
                    order.status
                ])
            
            # Adjust column widths based on content
            order_table = Table(order_data, colWidths=[96, 90, 180, 90, 70])
            order_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),  # Amount column right-aligned
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('WORDWRAP', (0, 1), (-1, -1), True), # Enable word wrapping for all columns
            ]))
            elements.append(order_table)
            
            # Build the PDF document
            doc.build(elements)
            pdf_data = buffer.getvalue()
            buffer.close()
            response.write(pdf_data)
            return response
        except Exception as e:
            # If PDF generation fails, return a simple error PDF
            buffer = BytesIO()
            p = canvas.Canvas(buffer, pagesize=A4)
            p.drawString(100, 750, "Sales Report")
            p.drawString(100, 700, "Error generating detailed report. Please try Excel format instead.")
            p.drawString(100, 680, f"Error: {str(e)[:100]}...")
            p.save()
            pdf_data = buffer.getvalue()
            buffer.close()
            response.write(pdf_data)
            return response

    elif report_format == 'excel':
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'

        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Report"

        # Header
        ws.append(['Sales Report'])
        ws.append([f'Period: {start_date} to {end_date}'])
        ws.append([])

        # Summary
        ws.append(['Summary'])
        ws.append(['Total Orders:', total_orders])
        ws.append(['Total Revenue:', f'₹{total_revenue:.2f}'])
        ws.append(['Units Sold:', total_units_sold])
        ws.append(['Unique Customers:', unique_customers])
        ws.append([])

        # Best-selling products
        ws.append(['Best Selling Products'])
        ws.append(['Product', 'Units Sold'])
        
        max_product_len = 0  # Track max product name length
        for product in best_selling_products:
            product_name = product['name']
            max_product_len = max(max_product_len, len(product_name))
            ws.append([product_name, product['sold_count']])
        ws.append([])

        # Category breakdown
        ws.append(['Sales by Category'])
        ws.append(['Category', 'Amount'])
        
        max_category_len = 0  # Track max category name length
        for category in category_data:
            max_category_len = max(max_category_len, len(category['name']))
            ws.append([category['name'], f'₹{category["amount"]:.2f}'])
        ws.append([])

        # Order details
        ws.append(['Order Details'])
        ws.append(['Order ID', 'Date', 'Customer', 'Products', 'Amount'])
        
        max_customer_len = 0
        max_products_len = 0
        
        for order in orders:
            items = OrderItem.objects.filter(order=order)
            product_str = ', '.join([f"{i.quantity} x {i.product.product_name}" for i in items])
            max_customer_len = max(max_customer_len, len(order.user.username))
            max_products_len = max(max_products_len, len(product_str))
            
            ws.append([
                order.invoice_no,
                order.date.strftime('%Y-%m-%d'),
                customer_name,
                product_str,
                f'₹{order.total_amount:.2f}',
                
            ])

        # Adjust column widths based on content
        # Column A - Order ID
        ws.column_dimensions['A'].width = 15
        # Column B - Date
        ws.column_dimensions['B'].width = 15
        # Column C - Customer (adjust based on max length)
        ws.column_dimensions['C'].width = min(max(max_customer_len + 2, 15), 50)
        # Column D - Products (adjust based on max length)
        ws.column_dimensions['D'].width = min(max(max_products_len + 2, 40), 100)
        # Column E - Amount
        ws.column_dimensions['E'].width = 15
        # Column F - Status
        ws.column_dimensions['F'].width = 15
        
        # Product column in the Best Selling Products section
        ws.column_dimensions['G'].width = min(max(max_product_len + 2, 40), 100)
        # Category column in the Sales by Category section
        ws.column_dimensions['H'].width = min(max(max_category_len + 2, 20), 50)
        
        wb.save(response)
        return response

    else:
        return redirect('inventory_dashboard')

@login_required
def order_details(request):
    from django.utils.timezone import now
    from django.http import JsonResponse
    
    products = tbl_product.objects.all()
    field_staff_list = FieldStaff.objects.select_related('userid').all()
    # Fetch all customers for the dropdown
    customer_list = Customer.objects.select_related('userid').all()
    current_date = now()
    invoice_no = f"INV-{current_date.strftime('%Y%m%d%H%M%S')}"
    
    # Handle AJAX request for checking product availability
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        product_id = request.POST.get('product_id')
        quantity = int(request.POST.get('quantity', 1))
        shade = request.POST.get('shade')
        
        try:
            product = tbl_product.objects.get(id=product_id)
            
            # Check availability based on product type
            if product.category.category_name.lower() == 'paints' and shade:
                if shade in product.available_shades:
                    available = product.available_shades.get(shade, 0)
                    if available < quantity:
                        return JsonResponse({
                            'available': False,
                            'message': f"Only {available} units of {product.product_name} in {shade} shade are available."
                        })
                    return JsonResponse({'available': True})
                else:
                    return JsonResponse({
                        'available': False,
                        'message': f"{product.product_name} is not available in {shade} shade."
                    })
            else:
                if product.quantity < quantity:
                    return JsonResponse({
                        'available': False,
                        'message': f"Only {product.quantity} units of {product.product_name} are available."
                    })
                return JsonResponse({'available': True})
                
        except tbl_product.DoesNotExist:
            return JsonResponse({'available': False, 'message': "Product not found."})
    
    # Handle normal POST request for order submission
    elif request.method == 'POST':
        try:
            date = request.POST.get('date')
            is_existing_customer = request.POST.get('is_existing_customer') == 'yes'
            customer_id = request.POST.get('customer_id')
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            address = request.POST.get('address')
            phone = request.POST.get('phone')
            approached_by_staff = request.POST.get('approached_by_staff')
            field_staff_id = request.POST.get('field_staff')
            product_ids = request.POST.getlist('product[]')
            quantities = request.POST.getlist('quantity[]')
            shades = request.POST.getlist('shade[]')
            
            # Validate required fields
            if not product_ids or all(not pid for pid in product_ids):
                messages.error(request, "Please add at least one product to the order.")
                return redirect('order_details')
                
            if not is_existing_customer and (not first_name or not last_name or not phone):
                messages.error(request, "Customer details are required for new customers.")
                return redirect('order_details')
            
            # Validate stock availability
            stock_error = None
            for i, product_id in enumerate(product_ids):
                if not product_id:  # Skip empty selections
                    continue
                    
                try:
                    product = tbl_product.objects.get(id=product_id)
                    quantity = int(quantities[i]) if quantities[i] else 1
                    
                    # Check paint products with shades
                    if product.category.category_name.lower() == 'paints' and i < len(shades) and shades[i]:
                        shade = shades[i]
                        # Check if shade exists and has sufficient stock
                        if shade in product.available_shades:
                            available_quantity = product.available_shades.get(shade, 0)
                            if available_quantity < quantity:
                                stock_error = f"Only {available_quantity} units of {product.product_name} in {shade} shade are available, but {quantity} were requested."
                                break
                            elif available_quantity == 0:
                                stock_error = f"{product.product_name} in {shade} shade is out of stock."
                                break
                        else:
                            stock_error = f"{product.product_name} is not available in {shade} shade."
                            break
                    else:
                        # Check regular products
                        if product.quantity < quantity:
                            stock_error = f"Only {product.quantity} units of {product.product_name} are available, but {quantity} were requested."
                            break
                        elif product.quantity == 0:
                            stock_error = f"{product.product_name} is out of stock."
                            break
                except tbl_product.DoesNotExist:
                    stock_error = f"Product not found for ID: {product_id}"
                    break
            
            if stock_error:
                messages.error(request, stock_error)
                return redirect('order_details')
            
            if approached_by_staff == 'yes' and not field_staff_id:
                messages.error(request, "Please select a field staff if approached by staff is 'Yes'.")
                return redirect('order_details')
            
            # Get the field staff if approached by staff
            field_staff = None
            if approached_by_staff == 'yes' and field_staff_id:
                try:
                    field_staff = get_object_or_404(FieldStaff, id=field_staff_id)
                except:
                    messages.error(request, "Selected field staff not found.")
                    return redirect('order_details')
            
            # Handle customer information
            customer = None
            if is_existing_customer and customer_id:
                try:
                    # Use existing customer
                    customer = Customer.objects.get(id=customer_id)
                    # Update approached_by_staff and staff_member if applicable
                    if approached_by_staff == 'yes':
                        customer.approached_by_staff = True
                        customer.staff_member = field_staff.userid  # This is correct as staff_member is a User
                    else:
                        customer.approached_by_staff = False
                    customer.save()
                except Customer.DoesNotExist:
                    messages.error(request, "Selected customer not found.")
                    return redirect('order_details')
            else:
                # Check if customer already exists based on phone number
                existing_customer = Customer.objects.filter(phone=phone).first()
                
                if not existing_customer:
                    try:
                        # Create a new user account for the customer
                        username = f"{first_name.lower()}_{phone[-4:]}"
                        # Check if username exists, if yes, modify it
                        count = 1
                        temp_username = username
                        while User.objects.filter(username=temp_username).exists():
                            temp_username = f"{username}_{count}"
                            count += 1
                        username = temp_username
                        
                        # Create user
                        user = User.objects.create_user(
                            username=username,
                            first_name=first_name,
                            last_name=last_name
                        )
                        
                        # Create customer
                        customer = Customer.objects.create(
                            userid=user,
                            address=address,
                            phone=phone,
                            customer_type='offline'  # Add missing required field
                        )
                        
                        # Set approached_by_staff and staff_member separately
                        customer.approached_by_staff = approached_by_staff == 'yes'
                        if field_staff and approached_by_staff == 'yes':
                            customer.staff_member = field_staff.userid  # This is correct as staff_member is a User
                        
                        customer.save()
                            
                    except Exception as e:
                        messages.error(request, f"Error creating customer: {str(e)}")
                        return redirect('order_details')
                else:
                    customer = existing_customer
                    # Update customer information
                    if customer.userid:
                        customer.userid.first_name = first_name
                        customer.userid.last_name = last_name
                        customer.userid.save()
                    
                    customer.address = address
                    
                    # Update staff-related fields
                    if approached_by_staff == 'yes':
                        customer.approached_by_staff = True
                        if field_staff:
                            customer.staff_member = field_staff.userid  # This is correct as staff_member is a User
                    
                    customer.save()
            
            # Create order
            try:
                # Use order date or current date
                order_date = date if date else current_date
                
                # Create order with the correct field_staff object
                order = Order.objects.create(
                    user=request.user,  # Adding the logged-in user as required by the model
                    customer=customer,
                    date=order_date,
                    invoice_no=invoice_no,
                    field_staff=field_staff,  # Use the FieldStaff object, not User
                    total_amount=0  # Initialize with 0, will update later
                )
                
                # Calculate total amount
                total_amount = 0
                
                # Create order items and update stock
                for i, product_id in enumerate(product_ids):
                    if not product_id:  # Skip empty selections
                        continue
                        
                    product = tbl_product.objects.get(id=product_id)
                    quantity = int(quantities[i]) if quantities[i] else 1
                    shade = shades[i] if i < len(shades) else None
                    
                    # Calculate item price
                    item_price = product.price * quantity
                    total_amount += item_price
                    
                    # Create order item
                    OrderItem.objects.create(
                        order=order,
                        product=product,
                        quantity=quantity,
                        color=shade if shade else '',  # Use color field as per model, not shade
                        subtotal=item_price  # Use subtotal field as per model, not price
                    )
                    
                    # Update product stock handled by the Order.save() method via signal
                
                # Update order total
                order.total_amount = total_amount
                order.status = 'completed' 
                order.save()
                
                messages.success(request, f"Order #{order.invoice_no} created successfully!")
                # Redirect to invoice page
                return redirect('order_invoice', order_id=order.id)
            except Exception as e:
                messages.error(request, f"Error creating order: {str(e)}")
                return redirect('order_details')
                
        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")
            return redirect('order_details')
    
    # For GET request, render the form
    context = {
        'products': products,
        'field_staff_list': field_staff_list,
        'customer_list': customer_list,
        'invoice_no': invoice_no,
        'current_date': current_date.strftime('%Y-%m-%d'),
    }
    return render(request, 'order_details.html', context)

def get_shades(request):
    product_id = request.GET.get('product_id')
    try:
        product = tbl_product.objects.get(id=product_id)
        if product.category.category_name.lower() == 'paints' and product.available_shades:
            # Extract shade names from the JSON field
            shades = list(product.available_shades.keys())
            return JsonResponse({'shades': shades})
    except tbl_product.DoesNotExist:
        return JsonResponse({'shades': []})
    return JsonResponse({'shades': []})

def order_invoice(request, order_id):
    try:
        order = Order.objects.get(id=order_id)
        order_items = OrderItem.objects.filter(order=order)
        
        context = {
            'order': order,
            'order_items': order_items,
            'customer': order.customer,
            'user': order.customer.userid,
            'field_staff': order.field_staff,
            'invoice_date': order.date,
            'invoice_no': order.invoice_no,
            'total_amount': order.total_amount,
        }
        
        # Generate PDF response if requested
        if 'print' in request.GET:
            template = get_template('invoice_view.html')
            html = template.render(context)
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="invoice_{order.invoice_no}.pdf"'
            
            # Create PDF from HTML content
            pisa_status = pisa.CreatePDF(html, dest=response)
            
            if pisa_status.err:
                return HttpResponse('Error generating PDF', status=500)
            return response
        
        # Add a print button to the template context
        context['show_print_button'] = True
        
        return render(request, 'invoice_view.html', context)
    
    except Order.DoesNotExist:
        messages.error(request, "Order not found.")
        return redirect('order_details')

from django.shortcuts import render
from django.http import JsonResponse
from ecom.models import tbl_category, tbl_product

from django.shortcuts import render, redirect
from ecom.models import tbl_category, tbl_product

def add_product(request):
    if request.method == "POST":
        product_name = request.POST.get('product_name')
        product_description = request.POST.get('product_description')
        price = request.POST.get('price')
        quantity = request.POST.get('quantity')
        category_id = request.POST.get('category')
        main_image = request.FILES.get('main_image')
        additional_images = {
            "additional_image_1": request.FILES.get('additional_image_1'),
            "additional_image_2": request.FILES.get('additional_image_2'),
            "additional_image_3": request.FILES.get('additional_image_3'),
        }

        category = tbl_category.objects.get(id=category_id)

        # Handling available shades (only for paint products)
        shades = {}
        if 'shadeColor[]' in request.POST and 'shadeQuantity[]' in request.POST:
            colors = request.POST.getlist('shadeColor[]')
            quantities = request.POST.getlist('shadeQuantity[]')
            shades = {colors[i]: int(quantities[i]) for i in range(len(colors))}

        # Create and save product
        product = tbl_product.objects.create(
            product_name=product_name,
            product_description=product_description,
            price=price,
            quantity=quantity,
            category=category,
            main_image=main_image,
            available_shades=shades if shades else None,
            **additional_images  # Assign additional images
        )

        messages.success(request, f"Product '{product_name}' has been added successfully.", extra_tags='add_product')
        return redirect('add_product')

    categories = tbl_category.objects.all()
    return render(request, 'addproduct.html', {'categories': categories})

@login_required
@user_passes_test(lambda u: ShowroomStaff.objects.filter(userid=u).exists())
def inventory(request):
    products = tbl_product.objects.all()
    categories = tbl_category.objects.all()
    
    # Apply filters
    search_query = request.GET.get('search', '')
    category_id = request.GET.get('category', '')
    stock_status = request.GET.get('stock', '')
    
    if search_query:
        products = products.filter(product_name__icontains=search_query)
    
    if category_id:
        products = products.filter(category_id=category_id)
    
    # Filter by stock status
    if stock_status == 'low':
        # Use a custom filter to find products with low stock
        low_stock_products = []
        for product in products:
            if 0 < product.total_stock < 5:
                low_stock_products.append(product.id)
        products = products.filter(id__in=low_stock_products)
    elif stock_status == 'out':
        # Use a custom filter to find products with zero stock
        out_of_stock_products = []
        for product in products:
            if product.total_stock == 0:
                out_of_stock_products.append(product.id)
        products = products.filter(id__in=out_of_stock_products)
    
    # Order the results
    products = products.order_by('category', 'product_name')
    
    context = {
        'products': products,
        'categories': categories,
    }
    return render(request, 'stockmanage.html', context)

@login_required
@user_passes_test(lambda u: ShowroomStaff.objects.filter(userid=u).exists())
def update_inventory(request, product_id):
    product = get_object_or_404(tbl_product, id=product_id)
    
    # Check for both shade-based method and 'paints' category
    is_shade_based = (product.category.quantity_calculation_method == "shade_based" or 
                      product.category.category_name.lower() == "paints")
    
    if request.method == 'POST':
        action_type = request.POST.get('action_type')
        
        if action_type == 'delete':
            # Get the product name before deletion for the success message
            product_name = product.product_name
            
            # Delete the product
            product.delete()
            
            # Show success message and redirect to inventory
            messages.success(request, f"Product '{product_name}' has been deleted successfully.", extra_tags='inventory')

            return redirect('inventory')
        if is_shade_based:

            # Handle shade-based products like paints
            shade_colors = request.POST.getlist('shade_color[]')
            shade_quantities = request.POST.getlist('shade_quantity[]')
            
            # Create a new shades dictionary with updated values
            updated_shades = {}
            for i in range(len(shade_colors)):
                if i < len(shade_quantities):
                    color = shade_colors[i]
                    quantity = int(shade_quantities[i])
                    updated_shades[color] = quantity
            
            # Update the product's available_shades
            product.available_shades = updated_shades
            
            # If category method is not set to shade_based but should be, fix it
            if product.category.quantity_calculation_method != "shade_based" and product.category.category_name.lower() == "paints":
                # Update the category's calculation method
                product.category.quantity_calculation_method = "shade_based"
                product.category.save()
            
            product.save()  # The save method will update the total quantity
        else:
            # Handle regular products
            quantity = int(request.POST.get('quantity', 0))
            product.quantity = quantity
            product.save()
        
        messages.success(request, f"Inventory for {product.product_name} has been updated successfully.")
        return redirect('inventory')  # Redirect to inventory list
    
    context = {
        'product': product,
        'is_shade_based': is_shade_based,
    }
    return render(request, 'update_inventory.html', context)
@user_passes_test(lambda u: ShowroomStaff.objects.filter(userid=u).exists())
def view_product(request, product_id):
    product = get_object_or_404(tbl_product, id=product_id)
    return render(request, 'view_products.html', {'product': product})
